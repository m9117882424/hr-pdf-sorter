from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
import pdfplumber
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageFilter, ImageOps

AH_START = 34  # AH
OUTPUT_HEADERS = [
    "Дата",
    "ИНН / Кимлик",
    "Фамилия / Наименование",
    "Имя",
    "Сумма",
    "Сумма 2",
    "Комментарий сверки",
]

MISMATCH_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
COMMENT_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")

DEFAULT_TESSERACT = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
if DEFAULT_TESSERACT.exists():
    pytesseract.pytesseract.tesseract_cmd = str(DEFAULT_TESSERACT)

DATE_RE = re.compile(r"\b\d{2}[./-]\d{2}[./-]\d{4}\b")
AMOUNT_RE = re.compile(r"\b\d{1,3}(?:[.\s]\d{3})*,\d{2}\b|\b\d+,\d{2}\b")


@dataclass
class PdfRecord:
    filename: str
    text_source: str = ""
    date: Optional[str] = None
    date_source: str = ""
    inn_kimlik: Optional[str] = None
    kimlik_source: str = ""
    surname_title: Optional[str] = None
    name: Optional[str] = None
    amounts: list[str] | None = None
    raw_top_date_ocr: str = ""
    raw_kimlik_ocr: str = ""
    raw_name_ocr: str = ""
    page_count: int = 0
    payment_pages: list[int] | None = None
    skipped_pages: list[int] | None = None

    def __post_init__(self) -> None:
        if self.amounts is None:
            self.amounts = []
        if self.payment_pages is None:
            self.payment_pages = []
        if self.skipped_pages is None:
            self.skipped_pages = []

    @property
    def full_name(self) -> str:
        return normalize_name(f"{self.surname_title or ''} {self.name or ''}")

    @property
    def reversed_full_name(self) -> str:
        return normalize_name(f"{self.name or ''} {self.surname_title or ''}")


def normalize_spaces(text) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def normalize_turkish_chars(text: str) -> str:
    table = str.maketrans({
        "İ": "I", "I": "I", "ı": "I", "i": "I",
        "Ş": "S", "ş": "S",
        "Ğ": "G", "ğ": "G",
        "Ü": "U", "ü": "U",
        "Ö": "O", "ö": "O",
        "Ç": "C", "ç": "C",
    })
    return str(text or "").translate(table)


def normalize_name(text) -> str:
    text = normalize_turkish_chars(str(text or "")).upper().replace("Ё", "Е")
    text = re.sub(r"[^А-ЯA-Z0-9 ]+", " ", text)
    return normalize_spaces(text)


def normalize_header(text) -> str:
    return normalize_spaces(str(text or "")).replace("Ё", "Е")


def normalize_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%d.%m.%Y")
    # Excel serial date fallback.
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        base = dt.datetime(1899, 12, 30)
        try:
            return (base + dt.timedelta(days=float(value))).strftime("%d.%m.%Y")
        except Exception:
            pass
    text = str(value).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        yyyy, mm, dd = m.groups()
        return f"{dd}.{mm}.{yyyy}"
    m = re.search(r"(\d{2})[./-](\d{2})[./-](\d{4})", text)
    if m:
        dd, mm, yyyy = m.groups()
        return f"{dd}.{mm}.{yyyy}"
    return normalize_spaces(text)


def normalize_id(value) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    # HR kimlik/TCKN must not start with zero. In OCR/Excel it can appear as
    # 12 digits with an extra leading zero; normalize it to the real value.
    if len(digits) > 10 and digits.startswith("0"):
        digits = digits.lstrip("0")
    return digits


def is_valid_kimlik(value: str) -> bool:
    digits = normalize_id(value)
    # For this HR workflow the compared kimlik is a personal TCKN: 11 digits
    # and it must not start with zero. If OCR returns 12 digits with an extra
    # leading zero, normalize_id() strips it before validation.
    return len(digits) == 11 and not digits.startswith("0")


def normalize_kimlik(value: str) -> str:
    digits = normalize_id(value)
    return digits if is_valid_kimlik(digits) else ""


def valid_kimlik_candidates(text: str) -> list[str]:
    candidates: list[str] = []
    # First take compact digit runs. This handles normal parsed PDF text.
    for m in re.finditer(r"\d{10,12}", str(text or "")):
        value = normalize_kimlik(m.group(0))
        if value and value not in candidates:
            candidates.append(value)
    # Then handle rare OCR spacing inside the number. Keep this conservative.
    for m in re.finditer(r"(?:\d\s*){10,12}", str(text or "")):
        value = normalize_kimlik(m.group(0))
        if value and value not in candidates:
            candidates.append(value)
    return candidates


def normalize_amount(text: str) -> str:
    text = re.sub(r"\s+", "", str(text or "").strip())
    m = re.search(
        r"\d{1,3}(?:[.\s]\d{3})*,\d{2}|\d+,\d{2}|\d{1,3}(?:[,\s]\d{3})*\.\d{2}|\d+\.\d{2}",
        text,
    )
    if not m:
        return text
    amount = m.group(0).replace(" ", "")
    if "," in amount and "." in amount:
        if amount.rfind(",") > amount.rfind("."):
            amount = amount.replace(".", "")
        else:
            amount = amount.replace(",", "")
    elif "," in amount:
        amount = amount.replace(".", "")
    return amount


def amount_to_excel_number(text: str):
    if text is None:
        return None
    text = re.sub(r"\s+", "", str(text).strip())
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def has_cyrillic(text: str) -> bool:
    return bool(re.search(r"[А-Яа-яЁё]", str(text or "")))


def valid_name_piece(text: str) -> bool:
    norm = normalize_name(text)
    if len(norm) < 2:
        return False
    if not re.search(r"[А-ЯA-Z]", norm):
        return False
    if re.fullmatch(r"[\d\s.:-]+", norm):
        return False
    bad_words = {
        "АДРЕС", "ADRES", "ADRESI", "ADRESI", "ДАТА", "TARIH", "ДАННЫЕ",
        "НАЛОГОВЫЙ", "НАЛОГОВАЯ", "ИНСПЕКЦИЯ", "VERGI", "DAIRESI",
        "ИМЯ", "AD", "ФАМИЛИЯ", "SOYADI", "ЗВАНИЕ", "UNVANI", "ДОЛЖНОСТЬ",
        "НАИМЕНОВАНИЕ", "НОМЕРНОЙ", "ЗНАК", "СЧЕТ", "HESAP", "ПОЛУЧАТЕЛЯ",
        "ОТПРАВИТЕЛЯ", "TOPLAM", "MIKTARI",
    }
    tokens = set(norm.split())
    if tokens and tokens.issubset(bad_words):
        return False
    return True


def cleanup_name_piece(text: str) -> str:
    text = normalize_spaces(text)
    text = re.sub(
        r"(?i)\b(?:Имя|Адрес|Ad[ıi]|Adresi|Adres|Plaka|Tarih|Vergi|Номерной\s+знак|Данные\s+о.*)$",
        "",
        text,
    ).strip()
    text = re.sub(r"^[\s:;|/\\-]+|[\s:;|/\\-]+$", "", text)
    text = re.sub(r"[^А-Яа-яЁёA-Za-z0-9İıŞşĞğÜüÖöÇç \-]+", " ", text)
    return normalize_spaces(text)


def check_tesseract() -> None:
    try:
        subprocess.run(
            [pytesseract.pytesseract.tesseract_cmd, "--version"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    except Exception as exc:
        raise RuntimeError(
            "Tesseract OCR не найден. Установите Tesseract для Windows и русский язык, "
            r"или проверьте путь C:\Program Files\Tesseract-OCR\tesseract.exe"
        ) from exc


def preprocess_for_ocr(img: Image.Image, scale: int = 2, threshold: bool = True) -> Image.Image:
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    if scale and scale > 1:
        img = img.resize((img.width * scale, img.height * scale), Image.Resampling.LANCZOS)
    img = img.filter(ImageFilter.SHARPEN)
    if threshold:
        img = img.point(lambda p: 255 if p > 175 else 0)
    return img


def ocr_image(
    img: Image.Image,
    *,
    lang: str = "rus+eng+tur",
    psm: int = 6,
    whitelist: str | None = None,
    scale: int = 2,
    threshold: bool = True,
) -> str:
    config = f"--oem 3 --psm {psm}"
    if whitelist:
        safe = str(whitelist).replace(chr(34), "").replace(chr(39), "").replace(" ", "")
        config += f" -c tessedit_char_whitelist={safe}"
    prepared = preprocess_for_ocr(img, scale=scale, threshold=threshold)
    for lang_try in (lang, "rus+eng", "eng"):
        try:
            return pytesseract.image_to_string(prepared, lang=lang_try, config=config)
        except (pytesseract.TesseractError, ValueError):
            continue
    return ""


def render_pdf_pages(pdf_path: Path, dpi: int = 250) -> list[Image.Image]:
    pages: list[Image.Image] = []
    doc = fitz.open(str(pdf_path))
    try:
        matrix = fitz.Matrix(dpi / 72, dpi / 72)
        for page in doc:
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            pages.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
    finally:
        doc.close()
    return pages


def crop_rel(img: Image.Image, box: tuple[float, float, float, float]) -> Image.Image:
    w, h = img.size
    left, top, right, bottom = box
    return img.crop((int(w * left), int(h * top), int(w * right), int(h * bottom)))


def dark_ratio(img: Image.Image, box: tuple[float, float, float, float]) -> float:
    gray = crop_rel(img, box).convert("L")
    hist = gray.histogram()
    return sum(hist[:180]) / max(1, gray.width * gray.height)


def looks_like_payment_page_text(text: str) -> bool:
    norm = normalize_name(text)
    keys = (
        "VERGI", "TAHSIL", "TCKNO", "VKN", "SOYADI", "UNVANI", "MIKTARI", "ZIRAAT",
        "ИНН", "КИМЛИК", "КВИТАНЦИЯ", "ПЛАТЕЖ", "ФАМИЛИЯ",
    )
    return sum(1 for key in keys if key in norm) >= 2


def looks_like_payment_page_image(img: Image.Image) -> bool:
    return dark_ratio(img, (0.04, 0.10, 0.96, 0.84)) >= 0.010


def extract_text_pages_from_pdf(pdf_path: Path) -> list[str]:
    chunks: list[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            chunks.append(page.extract_text() or "")
    return chunks


def extract_field(patterns: list[str], text: str, flags=re.IGNORECASE | re.MULTILINE) -> Optional[str]:
    for pattern in patterns:
        m = re.search(pattern, text, flags)
        if m:
            value = cleanup_name_piece(m.group(1))
            if value:
                return value
    return None


def extract_date_from_text(text: str) -> Optional[str]:
    patterns = [
        r"Tarih\s+No\s*\n\s*(\d{2}[./-]\d{2}[./-]\d{4})",
        r"\bTarih\b[^\n\r]*?\n[^\d\n\r]*(\d{2}[./-]\d{2}[./-]\d{4})",
        r"Ödeme\s+Tarihi.*?\n\s*\d{2}\s+(\d{2}[./-]\d{2}[./-]\d{4})",
        r"Дата\s+№\s*документа.*?\n.*?(\d{2}[./-]\d{2}[./-]\d{4})",
        r"Дата\s+Номер.*?\n.*?(\d{2}[./-]\d{2}[./-]\d{4})",
        r"Дата\s+№.*?\n.*?(\d{2}[./-]\d{2}[./-]\d{4})",
        r"Дата\s+оплаты.*?\n\s*02\s+(\d{2}[./-]\d{2}[./-]\d{4})",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            return normalize_date(m.group(1))
    m = DATE_RE.search(text)
    return normalize_date(m.group(0)) if m else None


def extract_kimlik_from_text(text: str) -> Optional[str]:
    text = str(text or "")
    priority_patterns = [
        r"TCKNO\s*/\s*VKN\s*:\s*([\d\s]{10,14})",
        r"TCKNO\s*/\s*VKN\s*([\d\s]{10,14})",
        r"Vergi\s+Kimlik\s+Numarası.*?TCKNO\s*/\s*VKN\s*:?\s*([\d\s]{10,14})",
    ]
    for pattern in priority_patterns:
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            value = normalize_kimlik(m.group(1))
            if value:
                return value

    fallback_patterns = [
        r"ИНН\s*/\s*Кимлик.*?\n\s*033202\s+([\d\s]{10,14})",
        r"ИНН\s*/\s*Кимлик.*?\n\s*\d{6}\s+([\d\s]{10,14})",
        r"Иден\.?\s*№\s*/\s*ИНН\s*[:\-]?\s*([\d\s]{10,14})",
        r"Идентиф\.?\s*номер\s+Тур\.?Респ\.?/\s*ИНН\s*[:\-]?\s*([\d\s]{10,14})",
        r"Идентификационный\s+номер\s+налогоплательщика.*?([\d\s]{10,14})",
        r"ИНН\s*[:\-]?\s*([\d\s]{10,14})",
        r"Кимлик\s*[:\-]?\s*([\d\s]{10,14})",
    ]
    for pattern in fallback_patterns:
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            value = normalize_kimlik(m.group(1))
            if value:
                return value

    candidates = valid_kimlik_candidates(text)
    return candidates[0] if candidates else None


def extract_name_from_text(text: str) -> tuple[str, str]:
    raw = text or ""
    surname_patterns = [
        r"Soyad[ıi]\s*/\s*[ÜU]nvan[ıi]\s*[:\-]?\s*([^\n\r:]+)",
        r"Soyad[ıi]\s*[:\-]?\s*([^\n\r:]+)",
        r"Фамилия\s*/\s*(?:Наименование|Назв\.?\s*имя|Должность|Звание)\s*[:\-]?\s*([^\n\r:]+)",
        r"Фамилия\s*[:\-]?\s*([^\n\r:]+)",
    ]
    name_patterns = [
        r"(?:^|\n|\r)\s*Ad[ıi]\s*[:\-]?\s*([^\n\r:]+)",
        r"\bAd[ıi]\s*[:\-]?\s*([^\n\r:]+)",
        r"(?:^|\n|\r)\s*Имя\s*[:\-]?\s*([^\n\r:]+)",
        r"\bИмя\s*[:\-]?\s*([^\n\r:]+)",
    ]
    surname = ""
    name = ""
    for pattern in surname_patterns:
        m = re.search(pattern, raw, re.IGNORECASE)
        if m:
            candidate = cleanup_name_piece(m.group(1))
            if valid_name_piece(candidate):
                surname = candidate
                break
    for pattern in name_patterns:
        m = re.search(pattern, raw, re.IGNORECASE)
        if m:
            candidate = cleanup_name_piece(m.group(1))
            if valid_name_piece(candidate):
                name = candidate
                break
    return surname, name


def extract_name_from_filename(filename: str) -> tuple[str, str]:
    stem = Path(filename).stem
    stem = re.sub(r"(?i)[_\s-]*(RU|RUS|TR)$", "", stem)
    stem = re.sub(r"^ПП\s+от\s+\d{2}[.]\d{2}[.]\d{4}\s+", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\d{5,}$", "", stem)
    stem = normalize_name(stem)
    if not has_cyrillic(stem) and not re.search(r"[A-Z]{2,}", stem):
        return "", ""
    stem = re.sub(r"\b\d{2}\s+\d{2}\s+\d{4}\b", " ", stem)
    stem = normalize_spaces(stem)
    parts = stem.split()
    if len(parts) >= 2:
        return parts[0], " ".join(parts[1:])
    if len(parts) == 1 and valid_name_piece(parts[0]):
        return parts[0], ""
    return "", ""


def extract_amounts_from_text(text: str) -> list[str]:
    found: list[str] = []
    for line in str(text or "").splitlines():
        line_norm = normalize_name(line)
        if (
            line_norm.startswith("05 ")
            or " TOPLAM " in f" {line_norm} "
            or line_norm.startswith("TOPLAM")
            or " ИТОГО " in f" {line_norm} "
            or line_norm.startswith("ИТОГО")
        ):
            tokens = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", line)
            if tokens:
                found.append(normalize_amount(tokens[-1]))

    preferred_patterns = [
        r"Общее\s+количество\s+платежей\s*:\s*\d+\s+([\d\s.]+,\d{2})",
        r"(?:Сумма|Размер)\s*[:\-]?\s*([\d\s.]+,\d{2})",
    ]
    for pattern in preferred_patterns:
        for m in re.finditer(pattern, text, re.IGNORECASE | re.DOTALL):
            found.append(normalize_amount(m.group(1)))

    if not found:
        for m in AMOUNT_RE.finditer(text):
            found.append(normalize_amount(m.group(0)))

    result: list[str] = []
    for value in found:
        num = amount_to_excel_number(value)
        if num is not None and num >= 100 and value not in result:
            result.append(value)
    return result


def page_indexes_for_text(text_pages: list[str]) -> tuple[list[int], list[int]]:
    payment_indexes = [i for i, text in enumerate(text_pages) if looks_like_payment_page_text(text)]
    if not payment_indexes:
        payment_indexes = [i for i, text in enumerate(text_pages) if normalize_spaces(text)]
    if not payment_indexes:
        payment_indexes = list(range(len(text_pages)))
    skipped = [i for i in range(len(text_pages)) if i not in payment_indexes]
    return payment_indexes, skipped


def extract_ocr_text_pages(pdf_path: Path) -> tuple[list[str], list[int], list[int]]:
    images = render_pdf_pages(pdf_path, dpi=250)
    payment_indexes = [i for i, img in enumerate(images) if looks_like_payment_page_image(img)]
    if not payment_indexes:
        payment_indexes = list(range(len(images)))
    text_pages = []
    for i in payment_indexes:
        text_pages.append(ocr_image(images[i], lang="rus+eng+tur", psm=6, scale=1, threshold=True))
    skipped = [i for i in range(len(images)) if i not in payment_indexes]
    return text_pages, payment_indexes, skipped


def parse_pdf(pdf_path: Path, use_ocr: bool = True) -> PdfRecord:
    try:
        text_pages = extract_text_pages_from_pdf(pdf_path)
    except Exception:
        text_pages = []
    page_count = len(text_pages)
    payment_indexes, skipped_indexes = page_indexes_for_text(text_pages) if text_pages else ([], [])
    selected_text_pages = [text_pages[i] for i in payment_indexes] if payment_indexes else text_pages
    text = "\n".join(selected_text_pages)
    text_source = "PDF_TEXT"

    date = extract_date_from_text(text)
    inn_kimlik = extract_kimlik_from_text(text)
    surname_title, name = extract_name_from_text(text)
    amounts = extract_amounts_from_text(text)

    needs_ocr = use_ocr and (
        len(normalize_spaces(text)) < 80
        or not date
        or not inn_kimlik
        or not surname_title
        or not name
        or not amounts
    )

    raw_ocr_text = ""
    if needs_ocr:
        ocr_text_pages, ocr_payment_indexes, ocr_skipped_indexes = extract_ocr_text_pages(pdf_path)
        ocr_text = "\n".join(ocr_text_pages)
        raw_ocr_text = ocr_text
        page_count = page_count or (len(ocr_payment_indexes) + len(ocr_skipped_indexes))
        payment_indexes = ocr_payment_indexes
        skipped_indexes = ocr_skipped_indexes
        text_source = "OCR_PAYMENT_PAGES" if len(normalize_spaces(text)) < 80 else "PDF_TEXT+OCR_PAYMENT_PAGES"
        date = date or extract_date_from_text(ocr_text)
        inn_kimlik = inn_kimlik or extract_kimlik_from_text(ocr_text)
        if not (surname_title and name):
            ocr_surname, ocr_name = extract_name_from_text(ocr_text)
            surname_title = surname_title or ocr_surname
            name = name or ocr_name
        if not amounts:
            amounts = extract_amounts_from_text(ocr_text)
        text = f"{text}\n{ocr_text}" if text else ocr_text

    if not surname_title or not name:
        text_surname, text_name = extract_name_from_text(text)
        surname_title = surname_title or text_surname
        name = name or text_name

    file_surname, file_name = extract_name_from_filename(pdf_path.name)
    if not surname_title or not valid_name_piece(surname_title):
        surname_title = file_surname
    if not name or not valid_name_piece(name):
        name = file_name

    page_count = page_count or len(text_pages)
    payment_pages = [i + 1 for i in payment_indexes] if payment_indexes else list(range(1, page_count + 1))
    skipped_pages = [i + 1 for i in skipped_indexes]

    return PdfRecord(
        filename=pdf_path.name,
        text_source=text_source,
        date=normalize_date(date) if date else None,
        date_source="TEXT_OR_OCR_TARIH" if date else "",
        inn_kimlik=normalize_kimlik(inn_kimlik) if inn_kimlik else None,
        kimlik_source="TEXT_OR_OCR_TCKNO_VKN" if inn_kimlik else "",
        surname_title=normalize_spaces(surname_title) if surname_title else None,
        name=normalize_spaces(name) if name else None,
        amounts=amounts,
        raw_top_date_ocr=raw_ocr_text[:2000],
        raw_kimlik_ocr=raw_ocr_text[:2000],
        raw_name_ocr=raw_ocr_text[:2000],
        page_count=page_count,
        payment_pages=payment_pages,
        skipped_pages=skipped_pages,
    )


def find_header_row(ws, required_headers: list[str] | None = None, max_scan_rows: int = 30) -> int:
    for row in range(1, max_scan_rows + 1):
        values = {normalize_header(ws.cell(row=row, column=col).value).lower() for col in range(1, ws.max_column + 1)}
        has_id = "y.t.c № кимлики" in values
        has_date = "дата оплаты" in values
        has_fio = "фио" in values or "фио рус" in values
        if has_id and has_date and has_fio:
            return row
    raise ValueError("Не найдены обязательные заголовки: ФИО/ФИО рус, Дата оплаты, Y.T.C № Кимлики")


def map_headers(ws, header_row: int) -> dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            headers[normalize_header(value)] = col
    return headers


def row_name_values(ws, row: int, name_cols: list[int]) -> list[str]:
    result = []
    for col in name_cols:
        name = normalize_name(ws.cell(row=row, column=col).value)
        if name and name not in result:
            result.append(name)
    return result


def build_excel_indexes(ws, header_row: int, headers: dict[str, int]) -> dict:
    date_col = headers.get("Дата оплаты")
    kimlik_col = headers.get("Y.T.C № Кимлики")
    name_cols = [headers[h] for h in ("ФИО", "ФИО рус") if headers.get(h)]
    if not date_col or not kimlik_col or not name_cols:
        raise ValueError("В Excel не найдены обязательные столбцы: ФИО/ФИО рус, Дата оплаты, Y.T.C № Кимлики")

    exact = {}
    by_date_id = defaultdict(list)
    by_name_date = defaultdict(list)
    by_name_id = defaultdict(list)
    by_name = defaultdict(list)

    for row in range(header_row + 1, ws.max_row + 1):
        pay_date = normalize_date(ws.cell(row=row, column=date_col).value)
        kimlik = normalize_kimlik(ws.cell(row=row, column=kimlik_col).value) or normalize_id(ws.cell(row=row, column=kimlik_col).value)
        names = row_name_values(ws, row, name_cols)

        if pay_date and kimlik:
            by_date_id[(pay_date, kimlik)].append(row)

        for fio in names:
            exact[(fio, pay_date, kimlik)] = row
            by_name_date[(fio, pay_date)].append(row)
            by_name_id[(fio, kimlik)].append(row)
            by_name[fio].append(row)

    return {
        "exact": exact,
        "by_date_id": by_date_id,
        "by_name_date": by_name_date,
        "by_name_id": by_name_id,
        "by_name": by_name,
    }


def record_name_variants(rec: PdfRecord) -> list[str]:
    values = []
    for value in (rec.full_name, rec.reversed_full_name):
        if value and value not in values:
            values.append(value)
    return values


def find_excel_row(indexes: dict, rec: PdfRecord) -> tuple[Optional[int], str]:
    names = record_name_variants(rec)
    date = normalize_date(rec.date or "")
    kimlik = normalize_kimlik(rec.inn_kimlik or "") or normalize_id(rec.inn_kimlik or "")

    if date and kimlik:
        for fio in names:
            row = indexes["exact"].get((fio, date, kimlik))
            if row:
                return row, "ФИО + дата из документа + кимлик"

        rows = indexes["by_date_id"].get((date, kimlik), [])
        if len(rows) == 1:
            return rows[0], "дата из документа + кимлик, строка уникальна"

    if date:
        for fio in names:
            rows = indexes["by_name_date"].get((fio, date), [])
            if len(rows) == 1:
                return rows[0], "ФИО + дата из документа"
    if kimlik:
        for fio in names:
            rows = indexes["by_name_id"].get((fio, kimlik), [])
            if len(rows) == 1:
                return rows[0], "ФИО + кимлик"
    for fio in names:
        rows = indexes["by_name"].get(fio, [])
        if len(rows) == 1:
            return rows[0], "только ФИО, сотрудник уникален в Excel"
    return None, "Нет безопасного совпадения"


def ensure_output_headers(ws, header_row: int) -> None:
    for i, header in enumerate(OUTPUT_HEADERS):
        ws.cell(row=header_row, column=AH_START + i).value = header


def excel_name_values(ws, row: int, headers: dict[str, int]) -> list[str]:
    names: list[str] = []
    for header in ("ФИО", "ФИО рус"):
        col = headers.get(header)
        if not col:
            continue
        value = normalize_name(ws.cell(row=row, column=col).value)
        if value and value not in names:
            names.append(value)
    return names


def get_excel_comparison(ws, row: int, record: PdfRecord, headers: dict[str, int] | None) -> dict:
    result = {
        "comments": [],
        "mismatch_fields": set(),
        "source_columns": set(),
    }
    if not headers:
        return result

    date_col = headers.get("Дата оплаты")
    kimlik_col = headers.get("Y.T.C № Кимлики")
    name_cols = [headers[h] for h in ("ФИО", "ФИО рус") if headers.get(h)]

    pdf_date = normalize_date(record.date or "")
    if date_col:
        excel_date = normalize_date(ws.cell(row=row, column=date_col).value)
        if pdf_date and excel_date and pdf_date != excel_date:
            result["comments"].append(f"Дата: PDF {pdf_date} ≠ Excel {excel_date}")
            result["mismatch_fields"].add("date")
            result["source_columns"].add(date_col)

    pdf_kimlik = normalize_kimlik(record.inn_kimlik or "") or normalize_id(record.inn_kimlik or "")
    if kimlik_col:
        excel_kimlik = normalize_kimlik(ws.cell(row=row, column=kimlik_col).value) or normalize_id(ws.cell(row=row, column=kimlik_col).value)
        if pdf_kimlik and excel_kimlik and pdf_kimlik != excel_kimlik:
            result["comments"].append(f"Кимлик: PDF {pdf_kimlik} ≠ Excel {excel_kimlik}")
            result["mismatch_fields"].add("kimlik")
            result["source_columns"].add(kimlik_col)

    pdf_names = record_name_variants(record)
    excel_names = excel_name_values(ws, row, headers)
    if pdf_names and excel_names and not any(name in excel_names for name in pdf_names):
        result["comments"].append(f"ФИО: PDF {' / '.join(pdf_names)} ≠ Excel {' / '.join(excel_names)}")
        result["mismatch_fields"].add("name")
        result["source_columns"].update(name_cols)
    elif not pdf_names and excel_names:
        result["comments"].append(f"ФИО не распознано; в Excel: {' / '.join(excel_names)}")
        result["mismatch_fields"].add("name")
        result["source_columns"].update(name_cols)

    return result


def apply_mismatch_highlight(ws, row: int, comparison: dict) -> None:
    mismatch_fields = comparison.get("mismatch_fields", set())
    if "date" in mismatch_fields:
        ws.cell(row=row, column=AH_START).fill = MISMATCH_FILL
    if "kimlik" in mismatch_fields:
        ws.cell(row=row, column=AH_START + 1).fill = MISMATCH_FILL
    if "name" in mismatch_fields:
        ws.cell(row=row, column=AH_START + 2).fill = MISMATCH_FILL
        ws.cell(row=row, column=AH_START + 3).fill = MISMATCH_FILL

    for col in comparison.get("source_columns", set()):
        ws.cell(row=row, column=col).fill = MISMATCH_FILL

    if comparison.get("comments"):
        ws.cell(row=row, column=AH_START + 6).fill = COMMENT_FILL


def make_comparison_comment(ws, row: int | None, record: PdfRecord, headers: dict[str, int] | None, match_mode: str = "") -> str:
    if not row:
        return "Нет безопасного совпадения с Excel"
    comments = get_excel_comparison(ws, row, record, headers).get("comments", [])
    if match_mode and match_mode not in ("ФИО + дата из документа + кимлик",):
        comments.append(f"Режим сопоставления: {match_mode}")
    return "; ".join(comments)


def write_record_to_row(ws, row: int, record: PdfRecord, headers: dict[str, int] | None = None, match_mode: str = "") -> str:
    ws.cell(row=row, column=AH_START).value = record.date
    kimlik_cell = ws.cell(row=row, column=AH_START + 1)
    kimlik_cell.value = normalize_kimlik(record.inn_kimlik or "") or normalize_id(record.inn_kimlik or "")
    kimlik_cell.number_format = "@"
    ws.cell(row=row, column=AH_START + 2).value = record.surname_title
    ws.cell(row=row, column=AH_START + 3).value = record.name
    amount_1 = amount_to_excel_number(record.amounts[0]) if len(record.amounts) >= 1 else None
    amount_2 = amount_to_excel_number(record.amounts[1]) if len(record.amounts) >= 2 else None
    for offset, amount in ((4, amount_1), (5, amount_2)):
        cell = ws.cell(row=row, column=AH_START + offset)
        cell.value = amount
        if amount is not None:
            cell.number_format = "#,##0.00"

    comparison = get_excel_comparison(ws, row, record, headers)
    comments = list(comparison.get("comments", []))
    if match_mode and match_mode not in ("ФИО + дата из документа + кимлик",):
        comments.append(f"Режим сопоставления: {match_mode}")
    comment_text = "; ".join(comments)
    ws.cell(row=row, column=AH_START + 6).value = comment_text
    apply_mismatch_highlight(ws, row, comparison)
    return comment_text


def save_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        rows = [{"status": "empty"}]
    fieldnames = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def unique_pdf_files(pdf_dir: Path) -> list[Path]:
    result = []
    seen = set()
    for path in sorted(pdf_dir.rglob("*")):
        if path.is_file() and path.suffix.lower() == ".pdf":
            key = str(path.resolve()).lower()
            if key not in seen:
                seen.add(key)
                result.append(path)
    return result


def process(excel_path: Path, pdf_dir: Path, sheet_name: Optional[str] = None, use_ocr: bool = True) -> None:
    if use_ocr:
        check_tesseract()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel не найден: {excel_path}")
    if not pdf_dir.exists() or not pdf_dir.is_dir():
        raise FileNotFoundError(f"Папка с PDF не найдена: {pdf_dir}")

    pdf_files = unique_pdf_files(pdf_dir)
    if not pdf_files:
        raise FileNotFoundError(f"В папке нет PDF-файлов: {pdf_dir}")

    print(f"Найдено PDF-файлов: {len(pdf_files)}", flush=True)
    print("Открываю Excel...", flush=True)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = find_header_row(ws)
    headers = map_headers(ws, header_row)
    indexes = build_excel_indexes(ws, header_row, headers)
    ensure_output_headers(ws, header_row)

    parsed_rows = []
    unmatched_rows = []
    debug_rows = []

    total_pdf = len(pdf_files)
    for idx, pdf_path in enumerate(pdf_files, start=1):
        print(f"[{idx}/{total_pdf}] Читаю PDF: {pdf_path.name}", flush=True)
        rec = parse_pdf(pdf_path, use_ocr=use_ocr)
        row, match_mode = find_excel_row(indexes, rec)

        comparison_comment = make_comparison_comment(ws, row, rec, headers, match_mode)

        row_info = {
            "pdf_file": rec.filename,
            "text_source": rec.text_source,
            "Страниц PDF": rec.page_count,
            "Полезные страницы": ",".join(map(str, rec.payment_pages or [])),
            "Пропущенные страницы": ",".join(map(str, rec.skipped_pages or [])),
            "Дата": rec.date or "",
            "date_source": rec.date_source,
            "ИНН / Кимлик": rec.inn_kimlik or "",
            "kimlik_source": rec.kimlik_source,
            "Фамилия / Наименование": rec.surname_title or "",
            "Имя": rec.name or "",
            "Сумма": rec.amounts[0] if len(rec.amounts) >= 1 else "",
            "Сумма 2": rec.amounts[1] if len(rec.amounts) >= 2 else "",
            "Все суммы": ";".join(rec.amounts or []),
            "Кол-во сумм": len(rec.amounts or []),
            "match_key_name": rec.full_name,
            "match_key_name_reversed": rec.reversed_full_name,
            "match_key_date": normalize_date(rec.date or ""),
            "match_key_id": normalize_kimlik(rec.inn_kimlik or "") or normalize_id(rec.inn_kimlik or ""),
            "match_mode": match_mode,
            "excel_row": row or "",
            "Комментарий сверки": comparison_comment,
        }
        parsed_rows.append(row_info)
        debug_rows.append(
            {
                "pdf_file": rec.filename,
                "raw_top_date_ocr": rec.raw_top_date_ocr,
                "raw_kimlik_ocr": rec.raw_kimlik_ocr,
                "raw_name_ocr": rec.raw_name_ocr,
                "payment_pages": row_info["Полезные страницы"],
                "skipped_pages": row_info["Пропущенные страницы"],
            }
        )

        if row:
            write_record_to_row(ws, row, rec, headers=headers, match_mode=match_mode)
        else:
            unmatched_rows.append(row_info)

    output_excel = excel_path.with_name(f"{excel_path.stem}_RU_decont_filled.xlsx")
    parsed_report = excel_path.with_name("RU_decont_parsed_report.csv")
    unmatched_report = excel_path.with_name("RU_decont_unmatched_report.csv")
    debug_report = excel_path.with_name("RU_decont_ocr_debug_report.csv")

    print("Сохраняю отчёты CSV...", flush=True)
    save_csv(parsed_report, parsed_rows)
    save_csv(unmatched_report, unmatched_rows)
    save_csv(debug_report, debug_rows)

    print("Сохраняю итоговый Excel... Не закрывайте окно.", flush=True)
    wb.save(output_excel)

    print("Готово.")
    print(f"PDF файлов обработано: {len(pdf_files)}")
    print(f"Совпадений записано в Excel: {len(pdf_files) - len(unmatched_rows)}")
    print(f"Несопоставленных PDF: {len(unmatched_rows)}")
    print(f"Результат сохранён: {output_excel}")
    print(f"Отчёт по распознанным PDF: {parsed_report}")
    print(f"Отчёт по несопоставленным PDF: {unmatched_report}")
    print(f"OCR debug отчёт: {debug_report}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Заполняет Excel данными из русских и турецких PDF-платёжек")
    parser.add_argument("--excel", required=True, help="Путь к исходному Excel-файлу")
    parser.add_argument("--pdf-dir", default="pdf_result", help="Папка с PDF-файлами")
    parser.add_argument("--sheet", default=None, help="Имя листа Excel, если нужен не активный лист")
    parser.add_argument("--no-ocr", action="store_true", help="Отключить OCR")
    args = parser.parse_args()
    process(Path(args.excel), Path(args.pdf_dir), args.sheet, use_ocr=not args.no_ocr)


if __name__ == "__main__":
    main()
